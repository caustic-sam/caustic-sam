import os
import httpx
from bs4 import BeautifulSoup
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
import random

# Configuration
LIMIT_DOWNLOADS = True  # Set True for only 5 random downloads, False for all
BASE_URL = "https://csrc.nist.gov"
SP_PAGE_URL = BASE_URL + "/publications/sp"
DOWNLOAD_DIR = os.path.expanduser("~/Documents/nist_downloads_http2")
LOG_FILE = os.path.join(DOWNLOAD_DIR, "nist_sp_download.log")
EXCEL_FILE = os.path.join(DOWNLOAD_DIR, "download_log.xlsx")

# Ensure the download directory exists before configuring logging
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# Set up logging
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console.setFormatter(formatter)
logging.getLogger().addHandler(console)

def setup_download_dir():
    """Ensures the download directory exists."""
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

def fetch_page(url):
    """Fetches and returns the HTML content of a webpage."""
    try:
        with httpx.Client(http2=True, timeout=10) as client:
            response = client.get(url, follow_redirects=True)
            response.raise_for_status()
            return response.text
    except Exception as e:
        logging.error(f"Error fetching {url}: {e}")
        return None

def extract_intermediate_links(html):
    """Extracts links to SP-series intermediate pages."""
    soup = BeautifulSoup(html, "html.parser")
    intermediate_links = {}
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if href.startswith("/pubs/sp/"):
            title = a_tag.get_text(strip=True).replace("/", "_") or "SP_Document"
            intermediate_links[title] = BASE_URL + href
    return intermediate_links

def extract_pdf_link(intermediate_page_url):
    """Navigates to an intermediate page and extracts the PDF download link."""
    html = fetch_page(intermediate_page_url)
    if not html:
        return None
    soup = BeautifulSoup(html, "html.parser")
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if href.lower().endswith(".pdf"):
            return href if href.startswith("http") else "https://nvlpubs.nist.gov" + href
    return None

def fetch_summary(intermediate_page_url):
    """Fetches the summary text from the SP intermediate page."""
    html = fetch_page(intermediate_page_url)
    if not html:
        return "N/A"
    soup = BeautifulSoup(html, "html.parser")
    summary_tag = soup.find("meta", {"name": "description"})
    return summary_tag["content"] if summary_tag else "No summary available."

def log_to_excel(title, status, size_mb, time_taken, summary):
    """Logs the title, size, download time, and summary to an Excel file."""
    if not os.path.exists(EXCEL_FILE):
        # Create a new Excel workbook if it doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Download Log"
        sheet.append(["Title", "Timestamp", "Size (MB)", "Time (s)", "Status", "Summary"])
        workbook.save(EXCEL_FILE)

    # Append the new log entry
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([
        title,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        f"{size_mb:.2f}",
        f"{time_taken:.2f}",
        status,
        summary
    ])
    workbook.save(EXCEL_FILE)

def download_pdf(client, title, url, intermediate_url):
    """Downloads a PDF with retries, logs size, time, and summary."""
    file_path = os.path.join(DOWNLOAD_DIR, f"{title}.pdf")
    start_time = time.time()
    size_mb = 0
    summary = fetch_summary(intermediate_url)  # Fetch the summary before downloading

    try:
        with client.stream("GET", url, follow_redirects=True, timeout=20) as response:
            response.raise_for_status()
            total_size = int(response.headers.get("content-length", 0))
            size_mb = total_size / (1024 * 1024)  # Convert to MB

            with open(file_path, "wb") as file, tqdm(
                desc=f"Downloading: {title}",
                total=total_size,
                unit="B",
                unit_scale=True
            ) as progress:
                for chunk in response.iter_bytes(256 * 1024):  # Using 256KB chunks instead of 128KB
                    file.write(chunk)
                    progress.update(len(chunk))

        time_taken = time.time() - start_time
        logging.info(f"Downloaded: {file_path} ({size_mb:.2f} MB in {time_taken:.2f}s)")
        log_to_excel(title, "Success", size_mb, time_taken, summary)

    except Exception as e:
        time_taken = time.time() - start_time
        logging.error(f"Failed to download {title}: {e}")
        log_to_excel(title, "Failed", size_mb, time_taken, summary)

def download_all_pdfs():
    """Main function to fetch links, extract PDFs, and download concurrently."""
    setup_download_dir()
    main_page_html = fetch_page(SP_PAGE_URL)
    if not main_page_html:
        logging.error("Failed to fetch main page. Exiting.")
        return

    intermediate_links = extract_intermediate_links(main_page_html)
    logging.info(f"Found {len(intermediate_links)} intermediate links to process.")

    if LIMIT_DOWNLOADS:
        all_titles = list(intermediate_links.keys())
        # Randomly pick 5 titles from the available links
        selected_titles = random.sample(all_titles, min(5, len(all_titles)))
        # Restrict intermediate_links to only those 5 selected items
        intermediate_links = {title: intermediate_links[title] for title in selected_titles}
        logging.info(f"LIMIT_DOWNLOADS is True. Only these 5 documents will be downloaded: {', '.join(selected_titles)}")

    # Double-check that we only have 5 (or fewer) intermediate_links now
    logging.info(f"Processing {len(intermediate_links)} documents after applying the limit.")

    with ThreadPoolExecutor(max_workers=10) as executor, httpx.Client(http2=True) as client:
        futures = []
        for title, intermediate_url in intermediate_links.items():
            pdf_url = extract_pdf_link(intermediate_url)
            if pdf_url:
                futures.append(executor.submit(download_pdf, client, title, pdf_url, intermediate_url))

        for future in as_completed(futures):
            pass

    logging.info("All downloads complete.")

if __name__ == "__main__":
    try:
        download_all_pdfs()
    except KeyboardInterrupt:
        logging.error("Script interrupted by user.")
        print("\nDownload interrupted. Exiting gracefully.")