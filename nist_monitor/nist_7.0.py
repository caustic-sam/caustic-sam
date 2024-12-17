import os
import asyncio
import httpx
import logging
import time
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook, load_workbook
import random

# Configuration
LIMIT_DOWNLOADS = True  # Set True for only 5 random downloads, False for all
BASE_URL = "https://csrc.nist.gov"
SP_PAGE_URL = BASE_URL + "/publications/sp"
DOWNLOAD_DIR = os.path.expanduser("~/Documents/nist_downloads_http2")
LOG_FILE = os.path.join(DOWNLOAD_DIR, "nist_sp_download.log")
EXCEL_FILE = os.path.join(DOWNLOAD_DIR, "download_log.xlsx")

# Ensure directories exist
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# Logging configuration
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console.setFormatter(formatter)
logging.getLogger().addHandler(console)

# Utility Functions
def log_to_excel(title, status, size_mb, time_taken, summary):
    """Logs metadata to an Excel file."""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Download Log"
        sheet.append(["Title", "Timestamp", "Size (MB)", "Time (s)", "Status", "Summary"])
        workbook.save(EXCEL_FILE)

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([
        title,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        f"{size_mb:.2f}",
        f"{time_taken:.2f}",
        status,
        summary
    ])
    workbook.save(EXCEL_FILE)

async def fetch_page(url):
    """Fetches HTML content asynchronously."""
    try:
        async with httpx.AsyncClient(http2=True, timeout=10) as client:
            response = await client.get(url, follow_redirects=True)
            response.raise_for_status()
            return response.text
    except Exception as e:
        logging.error(f"Error fetching {url}: {e}")
        return None

def extract_intermediate_links(html):
    """Extracts intermediate links to SP-series pages."""
    soup = BeautifulSoup(html, "html.parser")
    intermediate_links = {}
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if href.startswith("/pubs/sp/"):
            title = a_tag.get_text(strip=True).replace("/", "_") or "SP_Document"
            intermediate_links[title] = BASE_URL + href
    return intermediate_links

async def extract_pdf_link(intermediate_page_url):
    """Extracts the PDF link from an intermediate page."""
    html = await fetch_page(intermediate_page_url)
    if not html:
        return None
    soup = BeautifulSoup(html, "html.parser")
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if href.lower().endswith(".pdf"):
            return href if href.startswith("http") else "https://nvlpubs.nist.gov" + href
    return None

async def fetch_summary(intermediate_page_url):
    """Fetches a summary asynchronously."""
    html = await fetch_page(intermediate_page_url)
    if not html:
        return "N/A"
    soup = BeautifulSoup(html, "html.parser")
    summary_tag = soup.find("meta", {"name": "description"})
    return summary_tag["content"] if summary_tag else "No summary available."

async def async_download_pdf(title, pdf_url, intermediate_url):
    """Downloads a PDF asynchronously."""
    file_path = os.path.join(DOWNLOAD_DIR, f"{title}.pdf")
    start_time = time.time()
    size_mb = 0
    summary = await fetch_summary(intermediate_url)

    try:
        async with httpx.AsyncClient(http2=True, timeout=20) as client:
            response = await client.get(pdf_url, follow_redirects=True)
            response.raise_for_status()
            total_size = int(response.headers.get("content-length", 0))
            size_mb = total_size / (1024 * 1024)  # Convert to MB

            # Save file
            with open(file_path, "wb") as file:
                file.write(response.content)

        time_taken = time.time() - start_time
        logging.info(f"Downloaded: {file_path} ({size_mb:.2f} MB in {time_taken:.2f}s)")
        log_to_excel(title, "Success", size_mb, time_taken, summary)

    except Exception as e:
        time_taken = time.time() - start_time
        logging.error(f"Failed to download {title}: {e}")
        log_to_excel(title, "Failed", size_mb, time_taken, summary)

async def async_download_all_pdfs():
    """Main function to orchestrate downloads."""
    logging.info("Starting async download process...")
    html = await fetch_page(SP_PAGE_URL)
    if not html:
        logging.error("Failed to fetch main page. Exiting.")
        return

    intermediate_links = extract_intermediate_links(html)
    logging.info(f"Found {len(intermediate_links)} documents to process.")

    if LIMIT_DOWNLOADS:
        all_titles = list(intermediate_links.keys())
        selected_titles = random.sample(all_titles, min(5, len(all_titles)))
        intermediate_links = {title: intermediate_links[title] for title in selected_titles}
        logging.info(f"LIMIT_DOWNLOADS is True. Only these documents will be downloaded: {', '.join(selected_titles)}")

    tasks = []
    for title, intermediate_url in intermediate_links.items():
        pdf_url = await extract_pdf_link(intermediate_url)
        if pdf_url:
            tasks.append(async_download_pdf(title, pdf_url, intermediate_url))

    await asyncio.gather(*tasks)
    logging.info("All downloads complete.")

if __name__ == "__main__":
    try:
        asyncio.run(async_download_all_pdfs())
    except KeyboardInterrupt:
        logging.error("Script interrupted by user.")
        print("\nDownload interrupted. Exiting gracefully.")