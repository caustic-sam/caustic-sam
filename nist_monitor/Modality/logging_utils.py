import os
import logging
from openpyxl import Workbook, load_workbook
from datetime import datetime
from config import DOWNLOAD_DIR, LOG_FILE, EXCEL_FILE

def setup_logging():
    """Configures logging to file and console."""
    logging.basicConfig(
        filename=LOG_FILE,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    console.setFormatter(formatter)
    logging.getLogger().addHandler(console)

def setup_download_dir():
    """Ensures the download directory exists."""
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

def log_to_excel(title, status, size_mb, time_taken, summary):
    """Logs the title, size, download time, and summary to an Excel file."""
    if not os.path.exists(EXCEL_FILE):
        # Create a new Excel workbook if it doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Download Log"
        sheet.append(["Title", "Timestamp", "Size (MB)", "Time (s)", "Status", "Summary"])
        workbook.save(EXCEL_FILE)

    # Append the new log entry
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([
        title,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        f"{size_mb:.2f}",
        f"{time_taken:.2f}",
        status,
        summary
    ])
    workbook.save(EXCEL_FILE)